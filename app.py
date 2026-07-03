"""
Drawing QA Toolkit — Streamlit Application (v2)
=================================================
Three-mode drawing review tool for Autoliv / CET Thailand:

1. CET Revision Comparison — CV diff + OCR text diff + AI reconciliation
2. Prototype Completeness Check — checklist verification against drawing
3. Typo & Cross-View Consistency Check — spelling + mismatch detection

v2: PDF support, local AI (Ollama), OCR text diffing, Excel export, expanded UI.
"""
import streamlit as st
import cv2
import numpy as np
import json
import os
import io
import base64
from PIL import Image

from core.comparator import compare_drawings, crop_region

st.set_page_config(
    page_title="Drawing QA Toolkit",
    page_icon="📐",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------------------------
# Custom CSS for a polished look
# ---------------------------------------------------------------------------
st.markdown("""
<style>
    .stApp {
        background: linear-gradient(135deg, #0f0f1a 0%, #1a1a2e 50%, #16213e 100%);
    }
    .main .block-container {
        padding-top: 1.5rem;
    }
    h1 {
        background: linear-gradient(90deg, #00d4ff, #7b2ff7);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-weight: 800;
    }
    .stMetric {
        background: rgba(255,255,255,0.05);
        border-radius: 10px;
        padding: 10px;
        border: 1px solid rgba(255,255,255,0.1);
    }
    div[data-testid="stMetricValue"] {
        font-size: 1.8rem;
        font-weight: 700;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px 8px 0 0;
        padding: 8px 20px;
    }
    .css-1v0mbdj {
        border-radius: 12px;
    }
    div[data-testid="stSidebar"] {
        background: rgba(15, 15, 26, 0.95);
    }
    .status-badge {
        display: inline-block;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 0.85em;
        font-weight: 600;
    }
    .status-present { background: #1a4d2e; color: #4ade80; }
    .status-missing { background: #4d1a1a; color: #f87171; }
    .status-ambiguous { background: #4d3b1a; color: #fbbf24; }
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_image(file) -> np.ndarray:
    """Load an uploaded image file as BGR numpy array."""
    img = Image.open(file).convert("RGB")
    return cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR)


def load_drawing(file):
    """
    Universal loader: handles PDF, PNG, JPG, TIFF.
    Returns (list_of_bgr_images, file_type_str).
    """
    try:
        from core.pdf_handler import load_drawing_file
        return load_drawing_file(file)
    except ImportError:
        # Fallback if PyMuPDF not installed
        img = load_image(file)
        return [img], "image"


def cv2_to_display(img_bgr: np.ndarray) -> np.ndarray:
    return cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)


def img_to_base64(img_bgr: np.ndarray, max_width: int = 600) -> str:
    """Encode a BGR image as a base64 JPEG string for embedding in HTML."""
    # Resize if too large
    h, w = img_bgr.shape[:2]
    if w > max_width:
        scale = max_width / w
        img_bgr = cv2.resize(img_bgr, (max_width, int(h * scale)))
    rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
    pil = Image.fromarray(rgb)
    buf = io.BytesIO()
    pil.save(buf, format="JPEG", quality=90)
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def render_interactive_flagged_changes(discrepancies, master, aligned_revision):
    """
    Render an interactive flagged-changes viewer using Streamlit components.
    Each discrepancy is shown as a clickable card with zoomed before/after crops.
    """
    if not discrepancies:
        st.success("No discrepancies detected above the current sensitivity threshold.")
        return

    # Severity & category colors
    severity_emoji = {"minor": "🟡", "moderate": "🟠", "major": "🔴"}
    category_emoji = {"geometry": "📐", "dimension": "📏", "text": "📝", "annotation": "🏷️", "unknown": "❓"}

    # Let user pick which discrepancy to inspect
    options = [
        f"#{d.id}  {severity_emoji.get(d.severity, '')} {d.severity.upper()}  |  "
        f"{category_emoji.get(d.category, '')} {d.category}  |  {d.description[:60] if d.description else 'No description'}"
        for d in discrepancies
    ]

    st.markdown("**Click on any change below to see a zoomed before/after view of that region:**")

    # Show all discrepancies as expandable cards
    for i, d in enumerate(discrepancies):
        sev_color = {"minor": "#fbbf24", "moderate": "#f97316", "major": "#ef4444"}.get(d.severity, "#9ca3af")
        cat_color = {
            "geometry": "#ef4444", "dimension": "#f97316",
            "text": "#eab308", "annotation": "#06b6d4", "unknown": "#9ca3af"
        }.get(d.category, "#9ca3af")

        header = (
            f"{severity_emoji.get(d.severity, '')} **Change #{d.id}** — "
            f"**{d.severity.upper()}** {d.category} change"
        )
        if d.description:
            header += f"  \n`{d.description}`"

        with st.expander(header, expanded=(i == 0)):
            # Crop the region from both master and aligned revision
            master_crop = crop_region(master, d.bbox, context_factor=2.5, min_size=200)
            revision_crop = crop_region(aligned_revision, d.bbox, context_factor=2.5, min_size=200)

            col_before, col_after = st.columns(2)
            with col_before:
                st.markdown(f"**BEFORE (Master)**")
                st.image(cv2_to_display(master_crop), use_container_width=True)
            with col_after:
                st.markdown(f"**AFTER (Revision)**")
                st.image(cv2_to_display(revision_crop), use_container_width=True)

            # Info row
            info_cols = st.columns(3)
            info_cols[0].markdown(f"**Category:** `{d.category}`")
            info_cols[1].markdown(f"**Severity:** `{d.severity}`")
            info_cols[2].markdown(f"**Area:** `{d.area_px} px²`")


def render_curtain_slider(master_img: np.ndarray, revision_img: np.ndarray):
    """
    Renders an interactive swipe/curtain comparison slider inside a Streamlit component.
    """
    master_b64 = img_to_base64(master_img)
    revision_b64 = img_to_base64(revision_img)
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
    * {{box-sizing: border-box;}}
    body {{
        margin: 0;
        padding: 0;
        background: transparent;
        display: flex;
        justify-content: center;
        align-items: center;
        height: 100vh;
        overflow: hidden;
    }}
    .img-comp-container {{
        position: relative;
        width: 100%;
        height: 100%;
        max-width: 900px;
        aspect-ratio: {master_img.shape[1]} / {master_img.shape[0]};
        background-color: #1a1a2e;
        border: 2px solid rgba(255, 255, 255, 0.1);
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3);
    }}
    .img-comp-img {{
        position: absolute;
        width: 100%;
        height: 100%;
        overflow: hidden;
    }}
    .img-comp-img img {{
        display: block;
        width: 100%;
        height: 100%;
        object-fit: contain;
    }}
    .img-comp-overlay {{
        width: 50%;
        clip-path: inset(0 0 0 0);
        z-index: 2;
    }}
    /* The slider line */
    .img-comp-slider {{
        position: absolute;
        z-index: 9;
        cursor: ew-resize;
        width: 4px;
        height: 100%;
        background-color: #00d4ff;
        left: 50%;
        top: 0;
        transform: translateX(-50%);
        box-shadow: 0 0 10px rgba(0, 212, 255, 0.8);
    }}
    /* Slider button */
    .img-comp-button {{
        position: absolute;
        width: 40px;
        height: 40px;
        border-radius: 50%;
        background-color: #16213e;
        border: 3px solid #00d4ff;
        top: 50%;
        left: 50%;
        transform: translate(-50%, -50%);
        z-index: 10;
        display: flex;
        justify-content: center;
        align-items: center;
        color: #00d4ff;
        font-weight: bold;
        box-shadow: 0 0 15px rgba(0, 212, 255, 0.5);
        user-select: none;
    }}
    /* Range input covering the whole area */
    .img-comp-input {{
        position: absolute;
        width: 100%;
        height: 100%;
        opacity: 0;
        z-index: 11;
        cursor: ew-resize;
        margin: 0;
        padding: 0;
    }}
    </style>
    </head>
    <body>

    <div class="img-comp-container">
      <!-- Underneath: Aligned Revision -->
      <div class="img-comp-img">
        <img src="data:image/jpeg;base64,{revision_b64}">
      </div>
      <!-- Overlay: Master -->
      <div class="img-comp-img img-comp-overlay" id="overlay">
        <img src="data:image/jpeg;base64,{master_b64}">
      </div>
      <!-- Controls -->
      <div class="img-comp-slider" id="slider-line"></div>
      <div class="img-comp-button" id="slider-btn">↔</div>
      <input type="range" class="img-comp-input" id="range-input" min="0" max="100" value="50">
    </div>

    <script>
    const input = document.getElementById('range-input');
    const overlay = document.getElementById('overlay');
    const line = document.getElementById('slider-line');
    const btn = document.getElementById('slider-btn');

    input.addEventListener('input', (e) => {
        const val = e.target.value;
        // Update clip path of the overlay
        overlay.style.width = val + '%';
        line.style.left = val + '%';
        btn.style.left = val + '%';
    });
    </script>

    </body>
    </html>
    """
    import streamlit.components.v1 as components
    # Calculate aspect ratio height for components.html
    h, w = master_img.shape[:2]
    # Restrict max display width to 900px, scale height accordingly
    display_width = 900
    display_height = int((h / w) * display_width) + 40 # add safety padding
    
    components.html(html_content, height=display_height)


def export_to_excel(report_data: dict, filename: str = "report.xlsx") -> bytes:
    """Convert a report dict to Excel bytes for download."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

        row = 1
        for section, data in report_data.items():
            ws.cell(row=row, column=1, value=section.upper()).font = Font(bold=True, size=14)
            row += 1

            if isinstance(data, list) and data:
                # Write as table
                if isinstance(data[0], dict):
                    headers = list(data[0].keys())
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=h)
                        cell.font = header_font
                        cell.fill = header_fill
                    row += 1
                    for item in data:
                        for col, h in enumerate(headers, 1):
                            val = item.get(h, "")
                            if isinstance(val, (dict, list, tuple)):
                                val = str(val)
                            ws.cell(row=row, column=col, value=val)
                        row += 1
                else:
                    for item in data:
                        ws.cell(row=row, column=1, value=str(item))
                        row += 1
            elif isinstance(data, dict):
                for key, val in data.items():
                    ws.cell(row=row, column=1, value=str(key))
                    ws.cell(row=row, column=2, value=str(val))
                    row += 1
            else:
                ws.cell(row=row, column=1, value=str(data))
                row += 1
            row += 1

        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        return None


# ---------------------------------------------------------------------------
# AI Backend selection (sidebar)
# ---------------------------------------------------------------------------

def ai_backend_selector():
    """Render AI backend selection in sidebar. Returns (backend, api_key, vision_model)."""
    st.sidebar.markdown("---")
    st.sidebar.subheader("🤖 AI Backend")

    # Check Ollama availability
    ollama_ok = False
    try:
        from core.local_llm import check_ollama, list_models
        ollama_ok = check_ollama()
    except Exception:
        pass

    backend_options = ["Anthropic (Cloud)"]
    if ollama_ok:
        backend_options.insert(0, "Ollama (Local)")

    backend_choice = st.sidebar.radio(
        "AI Engine",
        backend_options,
        help="Ollama runs entirely on your machine. Anthropic sends images to the cloud.",
    )

    api_key = None
    vision_model = None

    if "Ollama" in backend_choice:
        backend = "ollama"
        models = list_models()
        vision_models = [m for m in models if any(v in m for v in ["llava", "bakllava", "moondream"])]
        if vision_models:
            vision_model = st.sidebar.selectbox("Vision Model", vision_models)
        else:
            st.sidebar.warning("No vision model found. Pull one with: `ollama pull llava:7b`")
            vision_model = "llava:7b"
    else:
        backend = "anthropic"
        api_key = st.sidebar.text_input(
            "Anthropic API key", type="password",
            value=os.environ.get("ANTHROPIC_API_KEY", ""),
            help="Get one at console.anthropic.com",
        )

    return backend, api_key, vision_model


# ---------------------------------------------------------------------------
# OCR status indicator
# ---------------------------------------------------------------------------

def show_ocr_status():
    """Show which OCR engines are available."""
    try:
        from core.ocr_engine import get_available_engines
        engines = get_available_engines()
        if engines:
            st.sidebar.success(f"OCR: {', '.join(engines)}")
        else:
            st.sidebar.warning("OCR: No engines found. Install Tesseract and/or EasyOCR.")
    except Exception:
        st.sidebar.warning("OCR module not available")


# ---------------------------------------------------------------------------
# Title
# ---------------------------------------------------------------------------

st.title("📐 Drawing QA Toolkit")
st.caption("AI-powered drawing review for CET / Autoliv Thailand — revision comparison, "
           "completeness checking, and consistency verification")

show_ocr_status()

mode = st.sidebar.radio(
    "📋 Check Mode",
    [
        "1. CET Revision Comparison",
        "2. Prototype Completeness Check",
        "3. Typo & Cross-View Consistency Check",
    ],
)

# =============================================================================
# MODE 1 — CET Revision Comparison
# =============================================================================
if mode.startswith("1"):
    st.header("CET Revision Comparison")
    st.caption("Verify requested changes were applied, and flag anything that wasn't requested. "
               "Supports PDF, PNG, JPG, TIFF.")

    with st.sidebar:
        st.subheader("📁 Inputs")
        use_sample = st.checkbox("Use sample drawings (no upload needed)", value=True)
        if not use_sample:
            master_file = st.file_uploader("Master Drawing",
                                           type=["png", "jpg", "jpeg", "tiff", "tif", "pdf"],
                                           key="m1_master")
            revision_file = st.file_uploader("New Revision",
                                             type=["png", "jpg", "jpeg", "tiff", "tif", "pdf"],
                                             key="m1_rev")
            if master_file and master_file.name.lower().endswith(".pdf"):
                try:
                    from core.pdf_handler import get_page_count
                    raw = master_file.read()
                    master_file.seek(0)
                    n_pages = get_page_count(raw)
                    if n_pages > 1:
                        master_page = st.number_input("Master page #", 1, n_pages, 1)
                    else:
                        master_page = 1
                except Exception:
                    master_page = 1
            else:
                master_page = 1

            if revision_file and revision_file.name.lower().endswith(".pdf"):
                try:
                    from core.pdf_handler import get_page_count
                    raw = revision_file.read()
                    revision_file.seek(0)
                    n_pages = get_page_count(raw)
                    if n_pages > 1:
                        revision_page = st.number_input("Revision page #", 1, n_pages, 1)
                    else:
                        revision_page = 1
                except Exception:
                    revision_page = 1
            else:
                revision_page = 1
        else:
            master_file, revision_file = None, None
            master_page, revision_page = 1, 1

        st.subheader("⚙️ Sensitivity")
        threshold = st.slider("Pixel difference threshold", 5, 100, 30)
        min_area = st.slider("Minimum region size (px²)", 50, 2000, 250)
        use_sift = st.checkbox("Use SIFT alignment (more robust)", value=True)
        use_ssim = st.checkbox("Use SSIM comparison", value=True)
        use_ocr_diff = st.checkbox("Enable OCR text diff", value=True,
                                    help="Extract text with OCR and show text-level changes")

        st.subheader("🔍 Change Request Verification (AI)")
        ai_enabled = st.checkbox("Reconcile diffs against requested changes", value=False)
        if ai_enabled:
            backend, api_key, vision_model = ai_backend_selector()
            requested_changes_text = st.text_area(
                "Requested changes (one per line)",
                value="Move the top-right bolt hole further right and down\n"
                      "Remove the keyway slot above the center bore\n"
                      "Add a new small hole to the left of center bore\n"
                      "Update the R25.4 dimension callout",
                height=140,
            )
        else:
            backend, api_key, vision_model = "anthropic", None, None
            requested_changes_text = ""

        run_btn = st.button("🚀 Run Comparison", type="primary", use_container_width=True)

    if run_btn:
        # Load images
        master_pdf_bytes = None
        revision_pdf_bytes = None
        if use_sample:
            master = cv2.imread("samples/master_drawing.png")
            revision = cv2.imread("samples/revised_drawing.png")
        else:
            if not master_file or not revision_file:
                st.error("Please upload both a master drawing and a revision, or check 'Use sample drawings'.")
                st.stop()

            # Handle PDF files
            master_pdf_bytes = None
            revision_pdf_bytes = None
            
            if master_file.name.lower().endswith(".pdf"):
                from core.pdf_handler import pdf_to_images
                master_pdf_bytes = master_file.read()
                master_file.seek(0)
                pages = pdf_to_images(master_pdf_bytes, pages=[master_page])
                master = pages[0].image
            else:
                master = load_image(master_file)

            if revision_file.name.lower().endswith(".pdf"):
                from core.pdf_handler import pdf_to_images
                revision_pdf_bytes = revision_file.read()
                revision_file.seek(0)
                pages = pdf_to_images(revision_pdf_bytes, pages=[revision_page])
                revision = pages[0].image
            else:
                revision = load_image(revision_file)

        # Run comparison
        with st.spinner("Aligning drawings and computing discrepancy map..."):
            result = compare_drawings(
                master, revision,
                threshold=threshold, min_area=min_area,
                use_sift=use_sift, use_ssim=use_ssim,
                use_ocr=use_ocr_diff,
            )

        # OCR-based structured text diff
        text_diff_result = None
        if use_ocr_diff:
            try:
                from core.pdf_handler import get_pdf_or_ocr_text, warp_text_blocks, extract_vector_text
                from core.structured_diff import compute_structured_diff
                from core.ocr_engine import OCRResult

                with st.spinner("Extracting text and computing text-level differences..."):
                    # Get master text (vector if PDF, else OCR)
                    master_ocr = get_pdf_or_ocr_text(master, master_pdf_bytes, page_number=master_page)
                    
                    # Get revision text
                    if revision_pdf_bytes is not None:
                        try:
                            raw_revision_blocks = extract_vector_text(revision_pdf_bytes, page_number=revision_page)
                            if raw_revision_blocks:
                                # Warp the bounding boxes of the vector text blocks using homography H
                                warped_blocks = warp_text_blocks(raw_revision_blocks, result.homography)
                                revision_ocr = OCRResult(
                                    text_blocks=warped_blocks,
                                    full_text=" ".join(b.text for b in warped_blocks),
                                    engine_used="pdf_vector_aligned",
                                    image_shape=result.aligned_revision.shape[:2]
                                )
                            else:
                                revision_ocr = get_pdf_or_ocr_text(result.aligned_revision)
                        except Exception:
                            revision_ocr = get_pdf_or_ocr_text(result.aligned_revision)
                    else:
                        revision_ocr = get_pdf_or_ocr_text(result.aligned_revision)

                    text_diff_result = compute_structured_diff(
                        master_ocr.text_blocks, revision_ocr.text_blocks
                    )
            except Exception as e:
                st.warning(f"Text diff could not be computed: {e}")

        # AI reconciliation
        reconciliation = None
        requested_changes = [c.strip() for c in requested_changes_text.splitlines() if c.strip()]
        if ai_enabled:
            if backend == "anthropic" and not api_key:
                st.error("Enter your Anthropic API key in the sidebar to run change-request verification.")
            elif not requested_changes:
                st.error("Enter at least one requested change in the sidebar.")
            elif not result.discrepancies:
                st.info("No pixel-level discrepancies were found, so there's nothing to reconcile.")
            else:
                from core.change_verifier import reconcile_changes
                with st.spinner(f"Reconciling {len(result.discrepancies)} region(s) against "
                                 f"{len(requested_changes)} requested change(s)..."):
                    try:
                        reconciliation = reconcile_changes(
                            master, result.aligned_revision, result.discrepancies,
                            requested_changes, backend=backend,
                            api_key=api_key, vision_model=vision_model,
                        )
                    except Exception as e:
                        st.error(f"AI reconciliation failed: {e}")

        # Title Block & BOM Analysis
        with st.spinner("Analyzing drawing layout (Title Block & BOM)..."):
            from core.layout_analyzer import parse_title_block_heuristics, parse_title_block_vlm, parse_bom_heuristics, parse_bom_vlm
            from core.pdf_handler import get_pdf_or_ocr_text

            if 'master_ocr' not in locals() or master_ocr is None:
                try:
                    master_ocr = get_pdf_or_ocr_text(master, master_pdf_bytes, page_number=master_page)
                except Exception:
                    from core.ocr_engine import extract_text
                    master_ocr = extract_text(master)

            tb_data = parse_title_block_heuristics(master_ocr.text_blocks, master.shape[:2])
            bom_data = parse_bom_heuristics(master_ocr.text_blocks, master.shape[:2])

            if ai_enabled:
                try:
                    vlm_tb = parse_title_block_vlm(master, backend=backend, api_key=api_key, vision_model=vision_model)
                    if vlm_tb:
                        tb_data.update({k: v for k, v in vlm_tb.items() if v is not None})
                except Exception:
                    pass

                try:
                    vlm_bom = parse_bom_vlm(master, backend=backend, api_key=api_key, vision_model=vision_model)
                    if vlm_bom:
                        bom_data = vlm_bom
                except Exception:
                    pass

        # --- Display Results ---
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Alignment", "✅ OK" if result.aligned_ok else "⚠️ FAILED")
        c2.metric("Match confidence", f"{result.match_confidence:.0%}")
        c3.metric("SSIM Score", f"{result.ssim_score:.3f}")
        c4.metric("Discrepancies", len(result.discrepancies))
        c5.metric("Processing time", f"{result.processing_time_s:.2f}s")

        if not result.aligned_ok:
            st.warning("⚠️ Automatic alignment had low confidence — results may include false positives from "
                       "scan skew/offset rather than real design changes.")

        # Build tabs
        tab_names = ["Side by Side", "Visual Diff (Red/Green)", "Interactive Swipe Slider", "Discrepancy Map", "Annotated Revision"]
        if text_diff_result:
            tab_names.append("Text Diff (OCR)")
        tab_names.append("Title Block & BOM")
        if reconciliation:
            tab_names.insert(0, "Change Verification")
        tab_names.append("Report")

        tabs = st.tabs(tab_names)
        tab_map = dict(zip(tab_names, tabs))

        if reconciliation:
            with tab_map["Change Verification"]:
                r = reconciliation.to_dict()
                cc1, cc2, cc3 = st.columns(3)
                cc1.metric("✅ Applied as requested", len(r["applied"]))
                cc2.metric("🚨 Unintended changes", len(r["unintended"]))
                cc3.metric("❓ Requested but missing", len(r["missing"]))
                if r["applied"]:
                    st.success("**Applied as requested**")
                    st.table(r["applied"])
                if r["unintended"]:
                    st.error("**Unintended changes** — flagged but not part of the request")
                    st.table(r["unintended"])
                if r["missing"]:
                    st.warning("**Requested but not found** — CET may not have applied these")
                    for m in r["missing"]:
                        st.write(f"- {m}")
                if not r["unintended"] and not r["missing"]:
                    st.success("✅ All detected changes match the request list, and nothing requested is missing.")

        with tab_map["Side by Side"]:
            st.image(cv2_to_display(result.side_by_side),
                     caption="Master (left) vs. Revision (right)",
                     use_container_width=True)

        with tab_map["Visual Diff (Red/Green)"]:
            if result.blended_diff is not None:
                st.image(cv2_to_display(result.blended_diff),
                         caption="Red/Green Blended Diff (Red = Deleted, Green = Added, Black = Unchanged)",
                         use_container_width=True)
            else:
                st.warning("Visual Diff not available")

        with tab_map["Interactive Swipe Slider"]:
            st.markdown("**Swipe the blue handle back and forth to compare drawings:**")
            render_curtain_slider(master, result.aligned_revision)

        with tab_map["Discrepancy Map"]:
            st.image(result.discrepancy_map,
                     caption="Binary discrepancy mask",
                     use_container_width=True)

        with tab_map["Annotated Revision"]:
            st.image(cv2_to_display(result.annotated_revision),
                     caption="Red=geometry, Orange=dimension, Yellow=text, Gray=unknown",
                     use_container_width=True)

        with tab_map["Title Block & BOM"]:
            st.subheader("📋 Title Block Metadata")
            col_tb1, col_tb2 = st.columns(2)
            with col_tb1:
                st.markdown(f"**Part / Drawing Number:** `{tb_data.get('part_number') or 'Not Found'}`")
                st.markdown(f"**Revision Level:** `{tb_data.get('revision') or 'Not Found'}`")
                st.markdown(f"**Title / Description:** `{tb_data.get('title') or 'Not Found'}`")
            with col_tb2:
                st.markdown(f"**Sheet Scale:** `{tb_data.get('scale') or 'Not Found'}`")
                st.markdown(f"**Measurement Units:** `{tb_data.get('units') or 'Not Found'}`")
                st.markdown(f"**Designer/Approver:** `{tb_data.get('designer') or 'Not Found'}`")

            st.subheader("📦 Bill of Materials (BOM)")
            if bom_data:
                import pandas as pd
                if isinstance(bom_data[0], dict) and "description" in bom_data[0]:
                    df_bom = pd.DataFrame(bom_data)
                    cols = [c for c in ["item", "part_number", "description", "qty", "material", "remarks"] if c in df_bom.columns]
                    st.dataframe(df_bom[cols] if cols else df_bom, use_container_width=True)
                else:
                    for item in bom_data:
                        st.markdown(f"**Item {item.get('item', '?')}:** {item.get('raw_row', '')}")
            else:
                st.info("No Bill of Materials (BOM) table detected or it is empty.")

        if text_diff_result and "Text Diff (OCR)" in tab_map:
            with tab_map["Text Diff (OCR)"]:
                summary = text_diff_result.summary()
                tc1, tc2, tc3, tc4 = st.columns(4)
                tc1.metric("Added text", summary["added"])
                tc2.metric("Removed text", summary["removed"])
                tc3.metric("Modified text", summary["modified"])
                tc4.metric("Moved text", summary["moved"])

                if text_diff_result.modified:
                    st.subheader("Modified Text")
                    mod_data = [
                        {"Master": c.master_text, "Revision": c.revision_text,
                         "Change": c.change_detail, "Severity": c.severity}
                        for c in text_diff_result.modified
                    ]
                    st.table(mod_data)

                if text_diff_result.added:
                    st.subheader("Added Text")
                    st.table([{"Text": c.revision_text, "Detail": c.change_detail}
                              for c in text_diff_result.added])

                if text_diff_result.removed:
                    st.subheader("Removed Text")
                    st.table([{"Text": c.master_text, "Detail": c.change_detail}
                              for c in text_diff_result.removed])

                if not text_diff_result.modified and not text_diff_result.added and not text_diff_result.removed:
                    st.success("No text-level differences detected by OCR.")

        with tab_map["Report"]:
            report = result.to_report_dict()
            if reconciliation:
                report["change_reconciliation"] = reconciliation.to_dict()
            if text_diff_result:
                report["text_diff"] = text_diff_result.to_report()

            col_json, col_excel = st.columns(2)
            with col_json:
                st.download_button("📥 Download JSON Report",
                                   data=json.dumps(report, indent=2),
                                   file_name="drawing_comparison_report.json",
                                   mime="application/json")
            with col_excel:
                excel_data = export_to_excel(report)
                if excel_data:
                    st.download_button("📥 Download Excel Report",
                                       data=excel_data,
                                       file_name="drawing_comparison_report.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            st.json(report)

        st.subheader("Flagged Changes — Interactive Viewer")
        render_interactive_flagged_changes(
            result.discrepancies, master, result.aligned_revision
        )
    else:
        st.info("Set your options in the sidebar and click **Run Comparison** to get started.")


# =============================================================================
# MODE 2 — Prototype Completeness Check
# =============================================================================
elif mode.startswith("2"):
    st.header("Prototype Completeness Check")
    st.caption("For prototype requests: confirm no missing or ambiguous instructions "
               "(base fabric type, sewing method, panel positioning method, etc.)")

    with st.sidebar:
        st.subheader("📁 Inputs")
        use_sample = st.checkbox("Use sample prototype drawing", value=True)
        drawing_page = 1
        if not use_sample:
            drawing_file = st.file_uploader("Prototype Drawing",
                                            type=["png", "jpg", "jpeg", "tiff", "tif", "pdf"],
                                            key="m2_drawing")
            if drawing_file and drawing_file.name.lower().endswith(".pdf"):
                try:
                    from core.pdf_handler import get_page_count
                    raw = drawing_file.read()
                    drawing_file.seek(0)
                    n_pages = get_page_count(raw)
                    if n_pages > 1:
                        drawing_page = st.number_input("Drawing page #", 1, n_pages, 1, key="m2_page")
                except Exception:
                    drawing_page = 1
        else:
            drawing_file = None

        st.subheader("📋 Required Instructions Checklist")
        from core.completeness_checker import load_checklist
        default_checklist = load_checklist()
        checklist_text = st.text_area(
            "One item per line — edit to match your drawing standard",
            value="\n".join(default_checklist),
            height=200,
        )

        backend, api_key, vision_model = ai_backend_selector()
        run_btn2 = st.button("🚀 Run Completeness Check", type="primary", use_container_width=True)

    if run_btn2:
        drawing_pdf_bytes = None
        if use_sample:
            drawing = cv2.imread("samples/prototype_drawing_sample.png")
            drawing_page = 1
        else:
            if not drawing_file:
                st.error("Upload a drawing or check 'Use sample prototype drawing'.")
                st.stop()

            if drawing_file.name.lower().endswith(".pdf"):
                from core.pdf_handler import pdf_to_images
                drawing_pdf_bytes = drawing_file.read()
                drawing_file.seek(0)
                pages = pdf_to_images(drawing_pdf_bytes, pages=[drawing_page])
                drawing = pages[0].image
            else:
                drawing = load_image(drawing_file)

        checklist = [c.strip() for c in checklist_text.splitlines() if c.strip()]

        if backend == "anthropic" and not api_key:
            st.error("Enter your Anthropic API key in the sidebar.")
        elif not checklist:
            st.error("Add at least one checklist item.")
        else:
            from core.completeness_checker import check_completeness
            with st.spinner(f"Checking drawing against {len(checklist)} required instruction(s)..."):
                try:
                    result = check_completeness(
                        drawing, checklist,
                        backend=backend, api_key=api_key,
                        vision_model=vision_model,
                        pdf_bytes=drawing_pdf_bytes,
                        page_number=drawing_page,
                    )
                except Exception as e:
                    st.error(f"Check failed: {e}")
                    result = None

            if result:
                r = result.to_dict()
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Checklist items", r["total_count"])
                c2.metric("✅ Present", r["present_count"])
                c3.metric("❌ Missing", r["missing_count"])
                c4.metric("⚠️ Ambiguous", r["ambiguous_count"])

                st.image(cv2_to_display(drawing), caption="Drawing under review",
                         use_container_width=True)

                for item in r["items"]:
                    if item["status"] == "present":
                        loc = f" (found in: {item['location']})" if item.get("location") else ""
                        st.success(f"**{item['item']}** — ✅ PRESENT{loc}\n\n> {item['evidence']}")
                    elif item["status"] == "missing":
                        st.error(f"**{item['item']}** — ❌ MISSING\n\n{item['note']}")
                    else:
                        st.warning(f"**{item['item']}** — ⚠️ AMBIGUOUS\n\n{item['note']}")

                # Downloads
                col_json, col_excel = st.columns(2)
                with col_json:
                    st.download_button("📥 Download JSON Report",
                                       data=json.dumps(r, indent=2),
                                       file_name="completeness_report.json",
                                       mime="application/json")
                with col_excel:
                    excel_data = export_to_excel(r)
                    if excel_data:
                        st.download_button("📥 Download Excel Report",
                                           data=excel_data,
                                           file_name="completeness_report.xlsx",
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("Set your checklist in the sidebar and click **Run Completeness Check** to get started.")


# =============================================================================
# MODE 3 — Typo & Cross-View Consistency Check
# =============================================================================
else:
    st.header("Typo & Cross-View Consistency Check")
    st.caption("Find spelling mistakes and mismatches between the main drawing and its views/sections.")

    with st.sidebar:
        st.subheader("📁 Inputs")
        use_sample = st.checkbox("Use sample drawing sheet", value=True)
        drawing_page = 1
        if not use_sample:
            drawing_file = st.file_uploader("Drawing (with views/sections)",
                                            type=["png", "jpg", "jpeg", "tiff", "tif", "pdf"],
                                            key="m3_drawing")
            if drawing_file and drawing_file.name.lower().endswith(".pdf"):
                try:
                    from core.pdf_handler import get_page_count
                    raw = drawing_file.read()
                    drawing_file.seek(0)
                    n_pages = get_page_count(raw)
                    if n_pages > 1:
                        drawing_page = st.number_input("Drawing page #", 1, n_pages, 1, key="m3_page")
                except Exception:
                    drawing_page = 1
        else:
            drawing_file = None

        st.subheader("📖 Known Domain Terms")
        from core.consistency_checker import load_domain_terms
        default_terms = load_domain_terms()
        domain_terms_text = st.text_area(
            "One term per line (won't be flagged as typos)",
            value="\n".join(default_terms),
            height=120,
        )

        backend, api_key, vision_model = ai_backend_selector()
        run_btn3 = st.button("🚀 Run Consistency Check", type="primary", use_container_width=True)

    if run_btn3:
        drawing_pdf_bytes = None
        if use_sample:
            drawing = cv2.imread("samples/prototype_drawing_sample.png")
            drawing_page = 1
        else:
            if not drawing_file:
                st.error("Upload a drawing or check 'Use sample drawing sheet'.")
                st.stop()

            if drawing_file.name.lower().endswith(".pdf"):
                from core.pdf_handler import pdf_to_images
                drawing_pdf_bytes = drawing_file.read()
                drawing_file.seek(0)
                pages = pdf_to_images(drawing_pdf_bytes, pages=[drawing_page])
                drawing = pages[0].image
            else:
                drawing = load_image(drawing_file)

        domain_terms = [t.strip() for t in domain_terms_text.splitlines() if t.strip()]

        if backend == "anthropic" and not api_key:
            st.error("Enter your Anthropic API key in the sidebar.")
        else:
            from core.consistency_checker import check_consistency
            with st.spinner("Reading all views and checking for typos and mismatches..."):
                try:
                    result = check_consistency(
                        drawing, domain_terms,
                        backend=backend, api_key=api_key,
                        vision_model=vision_model,
                        pdf_bytes=drawing_pdf_bytes,
                        page_number=drawing_page,
                    )
                except Exception as e:
                    st.error(f"Check failed: {e}")
                    result = None

            if result:
                r = result.to_dict()
                c1, c2 = st.columns(2)
                c1.metric("📝 Typos found", r["typo_count"])
                c2.metric("🔀 Cross-view mismatches", r["mismatch_count"])

                st.image(cv2_to_display(drawing), caption="Drawing under review",
                         use_container_width=True)

                if r["typos"]:
                    st.subheader("📝 Typos")
                    st.table(r["typos"])
                else:
                    st.success("✅ No typos found.")

                if r["mismatches"]:
                    st.subheader("🔀 Cross-View Mismatches")
                    st.table(r["mismatches"])
                else:
                    st.success("✅ No cross-view mismatches found.")

                # Downloads
                col_json, col_excel = st.columns(2)
                with col_json:
                    st.download_button("📥 Download JSON Report",
                                       data=json.dumps(r, indent=2),
                                       file_name="consistency_report.json",
                                       mime="application/json")
                with col_excel:
                    excel_data = export_to_excel(r)
                    if excel_data:
                        st.download_button("📥 Download Excel Report",
                                           data=excel_data,
                                           file_name="consistency_report.xlsx",
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("Click **Run Consistency Check** to get started.")
